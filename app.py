from flask import Flask, session, redirect, url_for, request, render_template, flash, make_response
from openpyxl import load_workbook
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
import os

app = Flask(__name__, static_folder='images', static_url_path='/images')
app.secret_key = 'super-secret-key-12345'
app.permanent_session_lifetime = timedelta(days=30)
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

DEFAULT_USERS = [
    {'username': 'joshua', 'email': 'joshua@gmail.com', 'password': 'password123'},
    {'username': 'admin', 'email': 'admin@example.com', 'password': 'admin123'}
]

DEFAULT_USER_SETTINGS = {
    'display_name': '',
    'role': 'Visitor',
    'likes': '',
    'dislikes': '',
    'birthday': '',
    'place': '',
    'fav_games': '',
    'tagline': '',
    'avatar_url': '/images/placeholder-image.png',
    'youtube_url': '',
    'tiktok_url': '',
    'facebook_url': '',
    'team_member_name': '',
    'team_member_pronouns': '',
    'team_member_role': '',
    'team_member_count': ''
}

def load_users_from_excel():
    excel_path = os.path.join(os.path.dirname(__file__), 'users.xlsx')
    if not os.path.exists(excel_path):
        return []

    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
    users = []
    for row in rows[1:]:
        user = {}
        for index, header in enumerate(headers):
            if header:
                user[header] = row[index] if index < len(row) else None
        users.append(user)
    return users


def user_exists_in_excel(username, email):
    users = load_users_from_excel()
    for user in users:
        if user.get('username', '').lower() == username.lower():
            return True, 'Username already exists'
        if user.get('email', '').lower() == email.lower():
            return True, 'Email already registered'
    return False, None


def add_user_to_excel(username, email, password, display_name='', role='Visitor'):
    excel_path = os.path.join(os.path.dirname(__file__), 'users.xlsx')
    if not os.path.exists(excel_path):
        return False, 'Excel file not found'
    
    try:
        workbook = load_workbook(excel_path)
        sheet = workbook.active
        
        next_row = sheet.max_row + 1
        sheet[f'A{next_row}'] = username
        sheet[f'B{next_row}'] = email
        sheet[f'C{next_row}'] = password
        sheet[f'D{next_row}'] = display_name
        sheet[f'E{next_row}'] = role
        sheet[f'Q{next_row}'] = datetime.now().strftime('%Y-%m-%d')
        
        workbook.save(excel_path)
        return True, 'User registered successfully'
    except Exception as e:
        return False, f'Error saving user: {str(e)}'


def update_user_in_excel(username, updates):
    excel_path = os.path.join(os.path.dirname(__file__), 'users.xlsx')
    if not os.path.exists(excel_path):
        return False, 'Excel file not found'

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    headers = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in sheet[1]]
    header_to_col = {header: index + 1 for index, header in enumerate(headers) if header}

    # Create any missing headers if needed
    for key in updates:
        if key not in header_to_col:
            new_col = len(headers) + 1
            sheet.cell(row=1, column=new_col).value = key
            headers.append(key)
            header_to_col[key] = new_col

    username_col = header_to_col.get('username')
    if username_col is None:
        return False, 'Username column not found in Excel file'

    row_to_update = None
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=username_col).value == username:
            row_to_update = row
            break

    if row_to_update is None:
        return False, 'User row not found in Excel file'

    for key, value in updates.items():
        col = header_to_col.get(key)
        if col:
            sheet.cell(row=row_to_update, column=col).value = value

    workbook.save(excel_path)
    return True, 'User settings updated'


def get_excel_user(user_input):
    if not user_input:
        return None

    users = load_users_from_excel()
    for user in users:
        if user.get('username') == user_input or user.get('email') == user_input:
            return user
    return None


def get_user_session():
    user = session.get('user', {})
    return {
        'username': session.get('username', user.get('username', 'Visitor')),
        'email': user.get('email', ''),
        'display_name': user.get('display_name') or user.get('username') or 'Visitor',
        'role': user.get('role') or 'Visitor',
        'likes': user.get('likes', ''),
        'dislikes': user.get('dislikes', ''),
        'birthday': user.get('birthday', ''),
        'place': user.get('place', ''),
        'fav_games': user.get('fav_games', ''),
        'tagline': user.get('tagline', ''),
        'avatar_url': user.get('avatar_url') or '/images/placeholder-image.png',
        'youtube_url': user.get('youtube_url', ''),
        'tiktok_url': user.get('tiktok_url', ''),
        'facebook_url': user.get('facebook_url', ''),
        'team_member_name': user.get('team_member_name', ''),
        'team_member_pronouns': user.get('team_member_pronouns', ''),
        'team_member_role': user.get('team_member_role', ''),
        'team_member_count': user.get('team_member_count', '')
    }

@app.route('/register', methods=['POST'])
def register():
    username = request.form.get('reg_username', '').strip()
    email = request.form.get('reg_email', '').strip()
    password = request.form.get('reg_password', '').strip()
    confirm_password = request.form.get('reg_confirm_password', '').strip()
    
    if not username or not email or not password or not confirm_password:
        return render_template('index.html', error='All fields are required')
    
    if password != confirm_password:
        return render_template('index.html', error='Passwords do not match')
    
    if len(password) < 6:
        return render_template('index.html', error='Password must be at least 6 characters')
    
    exists, error_msg = user_exists_in_excel(username, email)
    if exists:
        return render_template('index.html', error=error_msg)
    
    success, message = add_user_to_excel(username, email, password, display_name=username, role='Visitor')
    if not success:
        return render_template('index.html', error=message)
    
    session['username'] = username
    session['user'] = {
        'username': username,
        'email': email,
        **DEFAULT_USER_SETTINGS
    }
    print(f"DEBUG: New user registered: {username}")
    return redirect(url_for('homepage'))

@app.route('/', methods=['GET', 'POST'])
def auth():
    if request.method == 'POST':
        user_input = request.form.get('username_or_email', '').strip()
        password_input = request.form.get('password', '').strip()

        print(f"DEBUG: Login attempt for: {user_input}")

        # Check credentials from the Excel database
        user = get_excel_user(user_input)

        if user and user.get('password') == password_input:
            remember_me = request.form.get('remember_me') == 'on'
            session.permanent = remember_me
            session['username'] = user.get('username')
            session['user'] = {
                'username': user.get('username'),
                'email': user.get('email'),
                **DEFAULT_USER_SETTINGS,
                **{key: user.get(key) for key in DEFAULT_USER_SETTINGS.keys() if user.get(key)}
            }
            print(f"DEBUG: Login success! Session set. remember_me={remember_me}")
            response = make_response(redirect(url_for('homepage')))
            if remember_me:
                response.set_cookie('remembered_username', user_input, max_age=30*24*60*60)
            else:
                response.delete_cookie('remembered_username')
            return response
        else:
            print("DEBUG: Invalid credentials.")
            response = make_response(render_template('index.html', error='Invalid credentials', remembered_username=request.cookies.get('remembered_username', '')))
            return response
            
    remembered_username = request.cookies.get('remembered_username', '')
    return render_template('index.html', remembered_username=remembered_username)

@app.route('/homepage')
def homepage():
    print(f"DEBUG: Accessing homepage. Session: {session.get('username')}")
    if 'username' not in session:
        return redirect(url_for('auth'))
    return render_template('homepage.html', user=get_user_session())

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if 'username' not in session:
        return redirect(url_for('auth'))

    user = get_user_session()

    if request.method == 'POST':
        avatar_file = request.files.get('avatar_file')
        if avatar_file and avatar_file.filename:
            if allowed_file(avatar_file.filename):
                filename = secure_filename(avatar_file.filename)
                _, extension = os.path.splitext(filename)
                safe_filename = f"{session.get('username', 'user')}_{int(datetime.now().timestamp())}{extension}"
                save_path = os.path.join(os.path.dirname(__file__), 'images', safe_filename)
                avatar_file.save(save_path)
                user['avatar_url'] = f"/images/{safe_filename}"
            else:
                flash('Avatar must be a JPG or PNG file.', 'error')
                return redirect(url_for('settings'))

        editable_fields = ['display_name', 'likes', 'dislikes', 'birthday', 'place', 'fav_games', 'tagline', 'youtube_url', 'tiktok_url', 'facebook_url', 'team_member_name', 'team_member_pronouns', 'team_member_count']
        for field in editable_fields:
            user[field] = request.form.get(field, '').strip()

        session['user'] = user
        session['username'] = user['username']

        save_fields = ['display_name', 'role', 'likes', 'dislikes', 'birthday', 'place', 'fav_games', 'tagline', 'avatar_url', 'youtube_url', 'tiktok_url', 'facebook_url', 'team_member_name', 'team_member_pronouns', 'team_member_role', 'team_member_count']
        updates = {field: user.get(field, '') for field in save_fields}
        success, message = update_user_in_excel(user['username'], updates)
        if not success:
            flash(message, 'error')
            return redirect(url_for('settings'))

        flash('Settings saved successfully.', 'success')
        return redirect(url_for('settings'))

    return render_template('settings.html', user=user)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('auth'))

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=10000, debug=True)
